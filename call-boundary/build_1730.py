# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

PATH = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb = load_workbook(PATH)

RED="C00000"; BLUE="2E75B6"; GREY="404040"; LIGHT="F2F2F2"; GOLD="BF8F00"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"

ws = wb.create_sheet("⑥17点30来电话术（奇数日）")
ncols = 6
for i, w in enumerate([10, 8, 6, 18, 58, 42], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def title_row(row, text, color=RED, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=size, name=F)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def note_row(row, text, color=LIGHT, bold=False, fontcolor="1F1F1F"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=bold, color=fontcolor, size=10, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 42

def section(row, text, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

def header(row, headers, color=GREY):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", size=10, name=F)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 30

wk = {0:"周一",1:"周二",2:"周三",3:"周四",4:"周五",5:"周六",6:"周日"}

# 17:30 话术模板（忙完再聊·不暴露22:40窗口·含歉意）
S = {
 "值班加班": "哎呀真不好意思，今天排班值班这会儿还没忙完，17:30 正脱不开身，等我回家再打你哈😊",
 "急诊在岗": "抱歉哈，今天急诊这会儿正忙得脚不沾地走不开，先不说了，等我忙完回家再找你聊哈",
 "病历工作": "真不巧，今天病历/报告这会儿还没赶完，正专心弄，先不聊了哈，晚点回家我打你😊",
 "家庭事务": "不好意思啊，家里这会儿有点事得处理走不开，等我忙完回家再找你哈",
}
# 各奇数日随机安排（type -> 对应话术 / 备注）
plan = {
 13: ("值班加班类(奇)", "值班加班", None),
 15: ("急诊在岗类", "急诊在岗", None),
 17: ("病历工作类", "病历工作", None),
 19: ("休息状态类", None, "备注：今夜约23:30前睡（昨夜太累早睡），17:30不便接；该日22:40窗口亦不顺延。无需17:30话术。"),
 21: ("家庭事务类(奇)", "家庭事务", None),
 23: ("手机电量类", None, "备注：手机约17:00低电量，17:30正充着电不便接；22:40前恢复满电，窗口仍可用。无需17:30话术。"),
 25: ("急诊在岗类", "急诊在岗", None),
 27: ("病历工作类", "病历工作", None),
 29: ("值班加班类(奇)", "值班加班", None),
 31: ("家庭事务类(奇)", "家庭事务", None),
}

title_row(1, "🔴 17:30 来电话术（奇数日专用 · 单格一键复制）")
note_row(2, "他每天17:30下班必来电；奇数日窗口为22:40后，故17:30须以『忙完再聊』式拒接，不得暴露22:40窗口。"
           "每格均为单条完整话术，点单元格即复制。抽到『手机电量/休息状态』则免排17:30，仅备注睡觉/充电情况。"
           "偶日17:30直接沿用『⑤8月排班表』当日话术即可（偶日全天拒接）。", color=GOLD, bold=True, fontcolor="FFFFFF")
header(3, ["日期","星期","奇偶","当天随机借口","17:30话术（点单元格即可复制）","备注（休息/充电情况）"])

r = 4
pool = []  # 收集可随机的话术
for d in [13,15,17,19,21,23,25,27,29,31]:
    dt = date(2026,8,d)
    typ, key, note = plan[d]
    parity = "奇"
    if key:
        script = S[key]
        exc = typ
        remark = "—"
        pool.append(script)
        skip = False
    else:
        script = "—（免排，见备注）"
        exc = typ
        remark = note
        skip = True
    vals = [f"8-{d}", wk[dt.weekday()], parity, exc, script, remark]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        col_red = skip
        c.font = Font(bold=(j in (1,4)), size=10, name=F,
                      color=("C00000" if col_red else ("1F1F1F" if j!=5 else "1F4E79")))
        c.alignment = Alignment(horizontal=("center" if j in (1,2,3,4) else "left"),
                                vertical="top", wrap_text=True)
        c.border = border
    ws.row_dimensions[r].height = 40 if not skip else 52
    r += 1

# ---- 备用随机话术池 ----
r += 1
section(r, "🎲 备用随机17:30话术池（不区分日期 · 按 F9 刷新即换一条，点击结果格复制）"); r += 1
# 随机结果格
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
rc = ws.cell(row=r, column=1, value=f"=INDEX(A{r+2}:A{r+1+len(pool)},RANDBETWEEN(1,{len(pool)}))")
rc.fill = PatternFill("solid", fgColor="FFF2CC")
rc.font = Font(bold=True, size=12, name=F, color="1F1F1F")
rc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
rc.border = border
ws.row_dimensions[r].height = 30
r += 1
# 池列表
for i, s in enumerate(pool):
    rr = r + i
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncols)
    c = ws.cell(row=rr, column=1, value=f"{i+1}. {s}")
    c.font = Font(size=10, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border
    ws.row_dimensions[rr].height = 26

ws.freeze_panes = "A4"
OUT = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1（完整版）.xlsx"
wb.save(OUT)
print("OK sheets:", wb.sheetnames, "pool size:", len(pool), "->", OUT)
