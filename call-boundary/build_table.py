# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

RED   = "C00000"
BLUE  = "2E75B6"
GREY  = "404040"
LIGHT = "F2F2F2"
GOLD  = "BF8F00"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"

def title_row(ws, ncols, text, row=1, color=RED, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=size, name=F)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32

def note_row(ws, ncols, text, row, color=LIGHT, bold=False, size=10, fontcolor="1F1F1F"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=bold, color=fontcolor, size=size, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def section(ws, ncols, text, row, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

def header(ws, row, headers, color=GREY):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", size=10, name=F)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 22

def rows(ws, start, data, widths=None, heights=None, bold_first=False):
    r = start
    for rec in data:
        for j, v in enumerate(rec, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(bold=(bold_first and j == 1), size=10, name=F)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = border
        if heights and (r - start) < len(heights):
            ws.row_dimensions[r].height = heights[r - start]
        r += 1
    return r

# ============ Sheet 1 : 速查栏 ============
ws1 = wb.active
ws1.title = "①速查栏"
widths1 = [26, 18, 48]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
title_row(ws1, 3, "🔴 通话边界管控话术系统 V1.1 ｜ 高频速查栏")
note_row(ws1, 3, "来电时一眼可见 · 先扫这栏再决定动作", 2, color=GOLD, bold=True, fontcolor="FFFFFF")
header(ws1, 3, ["场景", "操作", "标准回复"])
data1 = [
    ["偶数日来电", "直接挂断 →", "今天家里有事，今天急诊，改日再聊😊"],
    ["奇数日 22:40 前来电", "直接挂断 →", "今日急诊加班，回家后我会打你电话"],
]
nr = rows(ws1, 4, data1, heights=[34, 34])
note_row(ws1, 3, "📌 唯一合法接听/回拨窗口：仅【奇数日 22:40 之后】、自身事务处理完毕时，可按需接听或回拨当日未接来电。", nr)
note_row(ws1, 3, "⏰ 所有接通的语音通话必须全程开启双闹钟（半程预警 + 终极挂断），严禁无时限畅聊。", nr + 1, color="FCE4D6")
ws1.freeze_panes = "A4"

# ============ Sheet 2 : 核心规则 ============
ws2 = wb.create_sheet("②核心规则")
widths2 = [30, 28, 46]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
title_row(ws2, 3, "二、核心规则（总则 + 分场景标准应答）")
section(ws2, 3, "一、核心总则", 2)
header(ws2, 3, ["条目", "内容", ""])
data2a = [
    ["1", "唯一接听/回拨窗口：仅奇数日 22:40 之后、自身事务处理完毕时，可按需接语音、回拨当日未接来电。", ""],
    ["2", "强制前置要求：所有接通的语音通话全程必须设置双闹钟（半程预警+终极挂断），严禁无时限畅聊。", ""],
    ["3", "主动呼叫禁令：复习/工作时段禁止主动拨打电话；回拨属可选项、非必须义务，主动权完全在己方。", ""],
]
r = rows(ws2, 4, data2a, heights=[42, 42, 42])
# hide the empty 3rd column visually by narrowing? keep for alignment; fill light
section(ws2, 3, "二、分场景标准应答", r)
header(ws2, r + 1, ["场景", "操作流程", "标准文字"])
data2b = [
    ["偶数日来电（非窗口，必拒）", "直接挂断电话 → 立即发送固定文字", "今天家里有事，急诊值班忙，改日再聊😊"],
    ["奇数日非窗口来电（22:40 前）", "直接挂断 → 发送固定文字", "今日急诊加班，回家后我回你电话"],
]
rows(ws2, r + 2, data2b, heights=[40, 40])

# ============ Sheet 3 : 借口库 ============
ws3 = wb.create_sheet("③借口库")
widths3 = [7, 22, 44, 32]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
title_row(ws3, 4, "三、随机轮换借口库（6款 · 可自动随机抽取）")
header(ws3, 2, ["序号", "借口类型", "标准话术", "适用场景"])
data3 = [
    ["1", "手机电量类", "手机快没电自动关机了，刚充上电才看到。", "非窗口未接来电、隔夜回复"],
    ["2", "值班加班类", "今天排班值班，要忙到晚上9点多，没空看手机。", "白天/傍晚邀约、非窗口来电解释"],
    ["3", "急诊在岗类", "在急诊科帮忙，手头事多，不方便接电话。", "任意时段临时来电，通用性最强"],
    ["4", "病历工作类", "在赶当天的病历和报告，得专心弄完，先不聊了。", "奇数日窗口超时收尾、非窗口消息推脱"],
    ["5", "家庭事务类（奇数日专用）", "家里有点事要帮忙处理，走不开。", "奇数日非窗口来电、窗口延长收尾"],
    ["6", "休息状态类", "昨晚太累很早就睡了，刚看到消息。", "隔夜回复未接来电/消息"],
]
rows(ws3, 3, data3, heights=[34] * 6)
# random draw
ws3.merge_cells(start_row=9, start_column=1, end_row=9, end_column=1)
c = ws3.cell(row=9, column=1, value="🎲 随机抽取结果（刷新即换）：")
c.font = Font(bold=True, size=11, name=F, color=RED)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws3.merge_cells(start_row=9, start_column=2, end_row=9, end_column=4)
fc = ws3.cell(row=9, column=2, value="=INDEX(C:C,RANDBETWEEN(3,8))")
fc.fill = PatternFill("solid", fgColor="FFF2CC")
fc.font = Font(bold=True, size=12, name=F, color="1F1F1F")
fc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
fc.border = border
ws3.row_dimensions[9].height = 30
note_row(ws3, 4, "📋 WPS 自动随机抽取设置方法（本表已内置公式，刷新即生效）：", 10, color=BLUE, bold=True, fontcolor="FFFFFF")
notes3 = [
    "1. 本表已按 A-D 列排好：序号 / 借口类型 / 标准话术 / 适用场景（话术在 C 列，数据行 3-8）。",
    "2. 随机抽取结果见上方 B9 单元格，公式为：=INDEX(C:C,RANDBETWEEN(3,8))",
    "3. 在 WPS 中按 F9（或重新打开/编辑表格）即可刷新，随机跳出一条话术，直接复制使用。",
    "4. 若你单独把 6 行话术粘贴成从 A1 开始（表头在 A1、数据 2-7 行），公式改为 =INDEX(C:C,RANDBETWEEN(2,7))。",
]
rr = 11
for n in notes3:
    note_row(ws3, 4, n, rr, color=LIGHT, size=10)
    rr += 1
ws3.freeze_panes = "A3"

# ============ Sheet 4 : 回拨·处罚·铁律 ============
ws4 = wb.create_sheet("④回拨·处罚·铁律")
widths4 = [18, 84]
for i, w in enumerate(widths4, start=1):
    ws4.column_dimensions[get_column_letter(i)].width = w
title_row(ws4, 2, "四~六、回拨规则 · 违规处罚 · 执行铁律")
section(ws4, 2, "四、未接来电回拨专项规则", 2)
section(ws4, 2, "1. 合法回拨前提", 3, color=GOLD)
header(ws4, 4, ["条件", "说明"])
data4a = [
    ["时间窗口", "仅奇数日 22:40 后、自身有空时执行"],
    ["骚扰拦截", "单日来电 ≥ 3 次的骚扰式呼叫，当日直接取消回拨资格"],
]
r = rows(ws4, 5, data4a, heights=[32, 32])
section(ws4, 2, "2. 标准回拨流程", r, color=GOLD)
header(ws4, r + 1, ["步骤", "操作要点"])
data4b = [
    ["① 间隔回拨", "不卡点立刻回拨，间隔 10~20 分钟再拨，强化「刚忙完才有空」人设"],
    ["② 统一开场", "刚忙完急诊/病历的事，才有空看手机，刚看到你之前打电话，找我有事吗？"],
    ["③ 双闹钟", "接通瞬间启动双闹钟（15分钟预警、25分钟终极挂断），严格卡时限"],
    ["④ 标准收尾", "不早了，我得洗漱休息了，明天还得早起交班，有事你打字留言就行。"],
]
r = rows(ws4, r + 2, data4b, heights=[34, 40, 34, 40])
section(ws4, 2, "3. 当日不回拨补处理", r, color=GOLD)
note_row(ws4, 2, "若没空或不想回拨，次日补发一句即可，无需解释细节：『昨晚值班/赶材料弄到太晚，看到的时候已经很晚了，就没回你，有事你打字说就行。』", r + 1, color="FCE4D6")
r = r + 2
section(ws4, 2, "五、违规处罚机制", r)
header(ws4, r + 1, ["项目", "内容"])
data4c = [
    ["违规判定", "非窗口时段接听语音 / 通话未设闹钟超时畅聊 / 非规定时段主动回拨，均算违规"],
    ["处罚标准", "单次违规 → 执行 2 次禁言：取消后续 2 期奇数日的语音接听资格，仅保留极简文字回复"],
    ["执行要求", "处罚期内所有语音一律拒接，用借口库话术推脱，不得中途破例解除"],
]
r = rows(ws4, r + 2, data4c, heights=[46, 46, 40])
section(ws4, 2, "六、执行铁律", r)
header(ws4, r + 1, ["序号", "铁律内容"])
data4d = [
    ["1", "所有借口点到为止，不展开细节、不辩解、不道歉，不给对方讨价还价的空间"],
    ["2", "借口随机轮换使用，避免固定套路被识破"],
    ["3", "通话全程必须开启倒计时闹钟，到点刚性挂断，不靠意志力硬扛"],
    ["4", "处罚一旦触发严格执行；一次心软破例，边界会全线崩溃"],
]
rows(ws4, r + 2, data4d, heights=[34, 34, 34, 34])

out = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb.save(out)
print("SAVED:", out)
