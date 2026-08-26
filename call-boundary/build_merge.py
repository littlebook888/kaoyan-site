# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

PATH = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb = load_workbook(PATH)

# 移除旧的 ⑤ 与 ⑥，重做合并详版
for name in ["⑤8月排班表", "⑥17点30来电话术（奇数日）"]:
    if name in wb.sheetnames:
        wb.remove(wb[name])

RED="C00000"; BLUE="2E75B6"; GREY="404040"; LIGHT="F2F2F2"; GOLD="BF8F00"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"

ws = wb.create_sheet("⑤8月排班表（含17点30专列）")
ncols = 9
for i, w in enumerate([10, 8, 6, 16, 12, 18, 48, 50, 34], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def title_row(row, text, color=RED, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=size, name=F)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32

def note_row(row, text, color=LIGHT, bold=False, fontcolor="1F1F1F"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=bold, color=fontcolor, size=10, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 48

def section(row, text, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

def header(row, headers, color=GREY):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", size=10, name=F)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 32

wk = {0:"周一",1:"周二",2:"周三",3:"周四",4:"周五",5:"周六",6:"周日"}
window_note = ("22:40后按需接听/回拨｜开场：刚忙完急诊/病历的事才有空看手机，刚看到你打电话，"
               "找我啥？｜收尾：不早了得洗漱，明天还得早起交班，有事你打字留言～｜全程双闹钟(15/25min)")
free_note = "今日畅聊（执行前）· 无限制，自由聊即可"

# 每日主数据：(mode, win, 推荐借口, 当日主话术, 17:30话术, 备注)
D = {
 12: ("畅聊(今日)","自由","—", free_note, "—（今日畅聊，正常接即可）", ""),
 13: ("拒接·跳过窗口","无","急诊/临时",
      "真不好意思啊，今天院里临时有事走不开，本来想好好聊的，实在抽不出空，先不说了哈，改天找你😊",
      "哎呀真不好意思，今天排班值班这会儿还没忙完，17:30正脱不开身，等我回家再打你哈😊", ""),
 14: ("拒接日","无","家庭+急诊",
      "哎呀抱歉，今天家里有点状况得处理，急诊也临时喊我，腾不出手，等忙完这阵再找你聊哈",
      "不好意思啊今天家里有点状况走不开，17:30这会儿正忙着，等空了找你哈😊", ""),
 15: ("窗口日","22:40后","急诊在岗", window_note,
      "抱歉哈，今天急诊这会儿正忙得脚不沾地走不开，先不说了，等我忙完回家再找你聊哈", ""),
 16: ("拒接日","无","值班加班",
      "不好意思今天排班值班，得弄到挺晚，先不说了哈，回家我再打给你",
      "哎呀真不好意思，今天排班值班这会儿还没忙完，17:30正脱不开身，等我回家再打你哈😊", ""),
 17: ("窗口日","22:40后","病历工作", window_note,
      "真不巧，今天病历/报告这会儿还没赶完，正专心弄，先不聊了哈，晚点回家我打你😊", ""),
 18: ("拒接日","无","病历报告",
      "真不巧，今天得赶完病历和报告，得专心弄，先不聊了，忙完找你哈😊",
      "真不巧今天病历/报告这会儿正赶着，17:30走不开，晚点我打你哈😊", ""),
 19: ("窗口日","22:40后","休息状态类", window_note,
      "—（免排，见备注）", "今夜约23:30前睡（昨夜太累早睡），17:30不便接；该日22:40窗口亦不顺延。"),
 20: ("拒接日","无","急诊在岗",
      "抱歉哈，今天院里临时叫帮忙，手头事一堆走不开，本来想聊的，实在没办法，改天哈",
      "抱歉哈，今天院里临时叫帮忙，这会儿正忙，17:30走不开，等我忙完找你哈😊", ""),
 21: ("窗口日","22:40后","家庭事务类(奇)", window_note,
      "不好意思啊，家里这会儿有点事得处理走不开，等我忙完回家再找你哈", ""),
 22: ("拒接日","无","疲惫休息",
      "真不好意思，今天有点累懵了刚看到消息，不太方便接，改日咱们好好聊😊",
      "真不好意思今天有点累懵了，17:30刚看到你电话，不太方便接，改日哈😊", ""),
 23: ("窗口日","22:40后","手机电量类", window_note,
      "—（免排，见备注）", "手机约17:00低电量，17:30正充着电不便接；22:40前恢复满电，窗口仍可用。"),
 24: ("拒接日","无","家庭+急诊",
      "哎呀今天家里有点事走不开，急诊也临时有事，腾不出手，先不说了哈，等空了找你",
      "不好意思啊今天家里有点事走不开，17:30这会儿正忙，等空了找你哈😊", ""),
 25: ("窗口日","22:40后","急诊在岗", window_note,
      "抱歉哈，今天急诊这会儿正忙得脚不沾地走不开，先不说了，等我忙完回家再找你聊哈", ""),
 26: ("拒接日","无","急诊在岗",
      "不好意思啊今天急诊忙得脚不沾地，走不开，本来想多聊会的，实在抽不出空，改天哈",
      "抱歉哈，今天急诊这会儿正忙得脚不沾地，17:30走不开，等我忙完找你哈😊", ""),
 27: ("窗口日","22:40后","病历工作", window_note,
      "真不巧，今天病历/报告这会儿还没赶完，正专心弄，先不聊了哈，晚点回家我打你😊", ""),
 28: ("拒接日","无","病历报告",
      "真不巧今天得赶材料，得专心弄完，先不聊了哈，忙完我打你😊",
      "真不巧今天得赶材料，这会儿正弄着，17:30走不开，忙完我打你哈😊", ""),
 29: ("窗口日","22:40后","值班加班", window_note,
      "哎呀真不好意思，今天排班值班这会儿还没忙完，17:30正脱不开身，等我回家再打你哈😊", ""),
 30: ("拒接日","无","值班加班",
      "抱歉哈今天值班走不开，得弄到挺晚，先不说了，回家我再回你电话",
      "抱歉哈今天值班走不开，17:30这会儿正忙着，回家我再回你电话哈😊", ""),
 31: ("窗口日","22:40后","家庭事务类(奇)", window_note,
      "不好意思啊，家里这会儿有点事得处理走不开，等我忙完回家再找你哈", ""),
}

title_row(1, "🔴 8月排班表（2026.08.13 起执行 · 含每日17:30专列 · 人化话术）")
note_row(2, "规则：今日(8-12)畅聊；明日(8-13)跳过窗口期不聊；奇日22:40后为接听窗口，偶日及跳过日均拒接。"
           "他每天17:30下班必来电——奇日17:30用『忙完再聊』拒接（不暴露22:40窗口）；偶日17:30沿用当日拒接话术。"
           "抽到『手机电量/休息状态』日免排17:30，仅备注睡觉/充电情况。所有话术人化、含歉意、有借口感。",
         color=GOLD, bold=True, fontcolor="FFFFFF")
header(3, ["日期","星期","奇偶","模式","接听窗口","推荐借口","当日主话术 / 要点","17:30来电话术（点单元格复制）","备注（休息/电量/睡觉）"])

r = 4
pool = []
for d in range(12, 32):
    dt = date(2026, 8, d)
    parity = "奇" if d % 2 == 1 else "偶"
    mode, win, exc, main, t17, remark = D[d]
    vals = [f"8-{d}", wk[dt.weekday()], parity, mode, win, exc, main, t17, remark if remark else "—"]
    is_skip = t17.startswith("—")
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        red = (mode.startswith("拒接")) or is_skip
        c.font = Font(bold=(j in (1,4)), size=10, name=F,
                      color=("C00000" if red else ("2E75B6" if mode=="窗口日" else "1F1F1F")))
        c.alignment = Alignment(horizontal=("center" if j in (1,2,3,4,5) else "left"),
                                vertical="top", wrap_text=True)
        c.border = border
    if not is_skip and t17 != "—（今日畅聊，正常接即可）":
        pool.append(t17)
    ws.row_dimensions[r].height = 50 if not is_skip else 40
    r += 1

# ---- 人化备选话术库 ----
r += 1
section(r, "💡 人化备选话术库（可随机替换，避免套路感 · 均含歉意与借口感）"); r += 1
alts = [
 ("拒接·急诊/值班向", "真不好意思，今天院里临时有事走不开，本来想聊的，实在抽不出空，先不说了哈😊"),
 ("拒接·急诊/值班向", "哎呀今天急诊忙得脚不沾地，走不开，等我忙完这阵再找你聊哈"),
 ("拒接·急诊/值班向", "抱歉哈今天排班值班得弄到挺晚，先不说了，回家我再打给你"),
 ("拒接·家庭/事务向", "不好意思啊今天家里有点状况得处理，腾不出手，等空了找你哈"),
 ("拒接·家庭/事务向", "真不巧家里有点事走不开，今天不太方便接，改日咱们好好聊😊"),
 ("拒接·病历/报告向", "真不巧今天得赶完病历和报告，得专心弄，先不聊了，忙完找你哈"),
 ("拒接·病历/报告向", "抱歉哈今天得赶材料，弄完再聊，先不说了哈😊"),
 ("拒接·疲惫/休息向", "真不好意思今天有点累懵了刚看到消息，不太方便接，改天哈"),
 ("拒接·疲惫/休息向", "哎呀今天状态不太行，刚看到你消息，先歇了，有事你打字留给我哈"),
 ("窗口日·开场", "刚忙完急诊/病历的事才有空看手机，刚看到你之前打电话，找我啥？"),
 ("窗口日·收尾", "不早了，我得洗漱休息了，明天还得早起交班，有事你打字留言就行～"),
]
for scene, txt in alts:
    c1 = ws.cell(row=r, column=1, value=scene)
    c1.font = Font(bold=True, size=10, name=F, color=BLUE)
    c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); c1.border = border
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
    c2 = ws.cell(row=r, column=2, value=txt)
    c2.font = Font(size=10, name=F)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); c2.border = border
    ws.row_dimensions[r].height = 28
    r += 1

# ---- 17:30 随机话术池 ----
r += 1
section(r, "🎲 17:30 随机话术池（不区分日期 · 按 F9 刷新即换一条，点结果格复制）"); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
# 结果格在 row r，池从第 r+1 行起、共 len(pool) 条，故范围为 A{r+1}:A{r+len(pool)}
rc = ws.cell(row=r, column=1, value=f"=INDEX(A{r+1}:A{r+len(pool)},RANDBETWEEN(1,{len(pool)}))")
rc.fill = PatternFill("solid", fgColor="FFF2CC")
rc.font = Font(bold=True, size=12, name=F, color="1F1F1F")
rc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); rc.border = border
ws.row_dimensions[r].height = 30
r += 1
for i, s in enumerate(pool):
    rr = r + i
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncols)
    c = ws.cell(row=rr, column=1, value=f"{i+1}. {s}")
    c.font = Font(size=10, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c.border = border
    ws.row_dimensions[rr].height = 26

ws.freeze_panes = "A4"
wb.save(PATH)
print("OK sheets:", wb.sheetnames, "pool:", len(pool))
