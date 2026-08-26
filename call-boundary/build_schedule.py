# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

PATH = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb = load_workbook(PATH)

RED = "C00000"; BLUE = "2E75B6"; GREY = "404040"; LIGHT = "F2F2F2"; GOLD = "BF8F00"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"

ws = wb.create_sheet("⑤8月排班表")
ncols = 7
for i, w in enumerate([10, 8, 6, 16, 12, 16, 62], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def title_row(row, text, color=RED, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=size, name=F)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def note_row(row, text, color=LIGHT, bold=False, fontcolor="1F1F1F"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=bold, color=fontcolor, size=10, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

def section(row, text, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

def header(row, headers, color=GREY):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", size=10, name=F)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 22

wk = {0:"周一",1:"周二",2:"周三",3:"周四",4:"周五",5:"周六",6:"周日"}

reject = {
 (2026,8,13): ("急诊/临时", "真不好意思啊，今天院里临时有事走不开，本来想好好聊的，实在抽不出空，先不说了哈，改天找你😊"),
 (2026,8,14): ("家庭+急诊", "哎呀抱歉，今天家里有点状况得处理，急诊也临时喊我，腾不出手，等忙完这阵再找你聊哈"),
 (2026,8,16): ("值班加班", "不好意思今天排班值班，得弄到挺晚，先不说了哈，回家我再打给你"),
 (2026,8,18): ("病历报告", "真不巧，今天得赶完病历和报告，得专心弄，先不聊了，忙完找你哈😊"),
 (2026,8,20): ("急诊在岗", "抱歉哈，今天院里临时叫帮忙，手头事一堆走不开，本来想聊的，实在没办法，改天哈"),
 (2026,8,22): ("疲惫休息", "真不好意思，今天有点累懵了刚看到消息，不太方便接，改日咱们好好聊😊"),
 (2026,8,24): ("家庭+急诊", "哎呀今天家里有点事走不开，急诊也临时有事，腾不出手，先不说了哈，等空了找你"),
 (2026,8,26): ("急诊在岗", "不好意思啊今天急诊忙得脚不沾地，走不开，本来想多聊会的，实在抽不出空，改天哈"),
 (2026,8,28): ("病历报告", "真不巧今天得赶材料，得专心弄完，先不聊了哈，忙完我打你😊"),
 (2026,8,30): ("值班加班", "抱歉哈今天值班走不开，得弄到挺晚，先不说了，回家我再回你电话"),
}
window_note = ("22:40后按需接听/回拨｜开场：刚忙完急诊/病历的事才有空看手机，刚看到你打电话，"
               "找我啥？｜收尾：不早了得洗漱，明天还得早起交班，有事你打字留言～｜全程双闹钟(15/25min)")
free_note = "今日畅聊（执行前）· 无限制，自由聊即可"

title_row(1, "🔴 8月排班表（2026.08.13 起执行 · 人化话术版）")
note_row(2, "规则：今日(8-12)畅聊；明日(8-13)跳过窗口期不聊；奇日 22:40 后为接听窗口，偶日及跳过日均拒接。"
           "话术已人化、暗含歉意、留「想聊但不得不」的借口感，可配合下方备选库轮换。", color=GOLD, bold=True, fontcolor="FFFFFF")
header(3, ["日期", "星期", "奇偶", "模式", "接听窗口", "推荐借口", "当日话术 / 要点"])

r = 4
for d in range(12, 32):
    dt = date(2026, 8, d)
    parity = "奇" if d % 2 == 1 else "偶"
    if d == 12:
        mode, win, exc, script = "畅聊(今日)", "自由", "—", free_note
    elif d == 13:
        mode, win, exc, script = "拒接·跳过窗口", "无", reject[(2026,8,13)][0], reject[(2026,8,13)][1]
    elif parity == "偶":
        mode, win, exc, script = "拒接日", "无", reject[(2026,8,d)][0], reject[(2026,8,d)][1]
    else:
        mode, win, exc, script = "窗口日", "22:40后", "—", window_note
    vals = [f"8-{d}", wk[dt.weekday()], parity, mode, win, exc, script]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(bold=(j in (1,4)), size=10, name=F,
                      color=("C00000" if mode.startswith("拒接") else ("2E75B6" if mode=="窗口日" else "1F1F1F")))
        c.alignment = Alignment(horizontal=("center" if j in (1,2,3,4,5) else "left"),
                                vertical="top", wrap_text=True)
        c.border = border
    ws.row_dimensions[r].height = 46 if mode != "窗口日" else 54
    r += 1

# ---- 人化备选话术库 ----
r += 1
section(r, "💡 人化备选话术库（可随机替换，避免套路感 · 均含歉意与借口感）"); r += 1
header(r, ["场景", "人化话术（含歉意·借口感）", "", "", "", "", ""]); r += 1
alts = [
 ("拒接·急诊/值班向", "真不好意思，今天院里临时有事走不开，本来想聊的，实在抽不出空，先不说了哈😊"),
 ("拒接·急诊/值班向", "哎呀今天急诊忙得脚不沾地，走不开，等我忙完这阵再找你聊哈"),
 ("拒接·急诊/值班向", "抱歉哈今天排班值班得弄到挺晚，先不说了，回家我再打给你"),
 ("拒接·家庭/事务向", "不好意思啊今天家里有点状况得处理，腾不出手，等空了找你哈"),
 ("拒接·家庭/事务向", "真不巧家里有点事走不开，今天不太方便接，改日咱们好好聊😊"),
 ("拒接·病历/报告向", "真不巧今天得赶完病历和报告，得专心弄，先不聊了，忙完找你哈"),
 ("拒接·病历/报告向", "抱歉哈今天得赶材料，弄完再聊，先不说了哈😊"),
 ("拒接·疲惫/休息向", "真不好意思今天有点累懵了刚看到消息，不太方便接，改天哈"),
 ("拒接·疲惫/休息向", "哎呀今天状态不太行，刚看到你消息，先歇了，有事你打字留给我哈"),
 ("窗口日·开场", "刚忙完急诊/病历的事才有空看手机，刚看到你之前打电话，找我啥？"),
 ("窗口日·收尾", "不早了，我得洗漱休息了，明天还得早起交班，有事你打字留言就行～"),
]
for scene, txt in alts:
    c1 = ws.cell(row=r, column=1, value=scene)
    c1.font = Font(bold=True, size=10, name=F, color=BLUE)
    c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c1.border = border
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    c2 = ws.cell(row=r, column=2, value=txt)
    c2.font = Font(size=10, name=F)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = border
    ws.row_dimensions[r].height = 30
    r += 1

ws.freeze_panes = "A4"
# move this sheet right after ④ for logical order (optional: keep at end)
wb.save(PATH)
print("OK rows:", r, "sheets:", wb.sheetnames)
