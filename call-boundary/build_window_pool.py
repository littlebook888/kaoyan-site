# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb = load_workbook(PATH)
ws = wb['⑤8月排班表（含17点30专列）']

BLUE="2E75B6"; LIGHT="F2F2F2"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"
ncols = 9

def section(row, text, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

# 从现有内容之后追加（跳过一行空）
r = ws.max_row + 2

# 窗口日·开场 随机话术池（接通瞬间，营造刚忙完才有空、想聊却刚得空的人化语气）
open_pool = [
 "刚忙完急诊/病历的事才有空看手机，刚看到你之前打电话，找我啥？",
 "哎呀今天可算忙完了，刚看到你之前打的电话，有啥事呀？",
 "不好意思哈，刚把手头活儿弄完才看到你来电，找我啥事？",
 "刚交完班/写好病历才有空摸手机，刚看到你电话，咋啦？",
 "今天科室的活儿刚收尾，我才得空看你消息，找我啥？",
 "刚把事儿都忙完，才看见你打的电话，有啥要跟我说的？",
]
# 窗口日·收尾 随机话术池（到点刚性挂断前，标准收尾，暗含"得睡、明早交班"）
close_pool = [
 "不早了，我得洗漱休息了，明天还得早起交班，有事你打字留言就行～",
 "哎呀都这个点了，我得去洗漱睡了，明早还得交班呢，有事先发我消息哈",
 "真不巧时间过这么快，我得准备睡了，明天一早要交班，有事你打字留给我",
 "不早啦，我得去洗漱了，明天还得早起，先不聊了哈，有事留言～",
 "得睡了，明天还得早起交班，就先到这吧，有啥你打字说就行",
 "时间不早了，我得休息了，明早交班得起得来，有事你发我就好哈",
]

def build_pool(start, title, pool):
    r = start
    section(r, f"🎲 {title}（按 F9 刷新即换一条，点结果格复制）"); r += 1
    # 结果格
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    rc = ws.cell(row=r, column=1,
                 value=f"=INDEX(A{r+1}:A{r+len(pool)},RANDBETWEEN(1,{len(pool)}))")
    rc.fill = PatternFill("solid", fgColor="FFF2CC")
    rc.font = Font(bold=True, size=12, name=F, color="1F1F1F")
    rc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); rc.border = border
    ws.row_dimensions[r].height = 30
    r += 1
    for i, s in enumerate(pool):
        rr = r + i
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncols)
        c = ws.cell(row=rr, column=1, value=f"{i+1}. {s}")
        c.font = Font(size=10, name=F)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c.border = border
        ws.row_dimensions[rr].height = 26
    return r + len(pool) + 1  # 下一段起始（含空一行）

r = build_pool(r, "窗口日·开场 随机话术池", open_pool)
r = build_pool(r, "窗口日·收尾 随机话术池", close_pool)

try:
    wb.save(PATH)
    print("OK saved. sheets:", wb.sheetnames, "| open_pool:", len(open_pool), "close_pool:", len(close_pool))
except PermissionError as e:
    print("PERMISSION_ERROR: 文件正被预览占用，请关闭WPS后重试。", e)
