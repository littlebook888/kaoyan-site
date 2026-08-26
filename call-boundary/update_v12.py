# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"D:\Agentwork\workbuddy\2026-08-12-22-11-36\通话边界管控话术系统V1.1.xlsx"
wb = load_workbook(PATH)
ws = wb['⑤8月排班表（含17点30专列）']

RED="C00000"; BLUE="2E75B6"; GOLD="BF8F00"; GREY="404040"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "微软雅黑"
ncols = 9

def setcell(rr, cc, val, color="1F1F1F", bold=False, center=False):
    c = ws.cell(rr, cc, value=val)
    c.font = Font(bold=bold, size=10, name=F, color=color)
    c.alignment = Alignment(horizontal=("center" if center else "left"),
                             vertical="top", wrap_text=True)
    c.border = border
    return c

# ---------- 1) 顶部 note 追加 8-20 新规 ----------
old_note = ws.cell(2,1).value or ""
new_note = (old_note +
    "｜⚠️8-20起规则升级：聊天频率改为「每周1次、最多2次」（自然周·周一~周日计），奇偶窗口制作废；"
    "聊天日(主聊日/可选加次)可全天聊(仍须双闹钟)，非聊天日全天拒接、17:30统一用『本周已聊/约好日子』理由。")
ws.cell(2,1).value = new_note
ws.cell(2,1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 64

# ---------- 2) 重写 8-20(行12) ~ 8-31(行23) 为周频率逻辑 ----------
# 字段顺序: (模式, 窗口, 借口, 主话术, 17:30话术, 备注)
N = {
 12: ("周频率·聊天日①","全天可聊(约定)","新规过渡",
      "今天刚好有空，咱们好好聊会儿～不过我得定个闹钟，待会还有事要弄😊",
      "来啦，今天约好聊的，正好这会儿有空～",
      "第1周(8-17~8-23)第1次｜8-20新规生效·电话隔离见效"),
 13: ("周频率·拒接(额度已用)","无(本周已聊)","周频率额度",
      "真不好意思啊，这周咱们刚聊过，今天手头事一堆走不开，先不说了哈，等约好的日子再好好聊😊",
      "不好意思哈，这周刚聊过，今天正忙着呢，先不说了，等约好的日子再聊哈😊",
      "第1周·额度已用"),
 14: ("周频率·拒接(额度已用)","无(本周已聊)","周频率额度",
      "哎呀今天不太方便接，这周咱们刚聊过，先不说了哈，等约好的日子😊",
      "真不好意思今天走不开，这周刚聊过，先不说了哈，等约好的日子再聊😊",
      "第1周·额度已用"),
 15: ("周频率·可选加次②","可聊(可选)","周频率·加次",
      "今天周日有空的话可以再聊会儿，不过我得定闹钟哈～不想聊也完全ok，你留言就行😊",
      "今天周日有空可以聊，来啦～（可选；若选休息则17:30免排）",
      "第1周·可选第2次｜原手机电量日：约17:00低电量充电，选休息则17:30免排"),
 16: ("周频率·拒接(新周)","无(本周未聊)","周频率·未聊",
      "新的一周啦，今天手头事多走不开，这周咱们约好的日子再聊哈😊",
      "不好意思哈今天走不开，这周还没聊呢，等咱约好的日子哈😊",
      "第2周(8-24~8-30)开始·未聊"),
 17: ("周频率·拒接(未聊)","无(本周未聊)","周频率·未聊",
      "真不好意思今天不太方便接，这周咱们约好的日子再聊哈😊",
      "哎呀今天走不开，这周还没聊呢，等约好的日子哈😊",
      "第2周·未聊"),
 18: ("周频率·拒接(未聊)","无(本周未聊)","周频率·未聊",
      "今天手头事多，先不说了哈，这周约好的日子再聊😊",
      "不好意思哈今天忙着，这周约好的日子再聊哈😊",
      "第2周·未聊"),
 19: ("周频率·可选加次②","可聊(可选)","周频率·加次",
      "今天周四有空可以加聊一次，想聊就聊、不想也行哈，记得定闹钟😊",
      "周四有空可以聊，来啦（可选）～",
      "第2周·可选第2次"),
 20: ("周频率·拒接(未聊)","无(本周未聊)","周频率·未聊",
      "今天不太方便接，这周咱们约好的日子(周日)再聊哈😊",
      "不好意思哈今天走不开，这周约好的日子再聊哈😊",
      "第2周·未聊(主聊日8-30)"),
 21: ("周频率·拒接(未聊)","无(本周未聊)","周频率·未聊",
      "今天手头事多走不开，这周约好的日子再聊哈😊",
      "哎呀今天走不开，这周约好的日子再聊哈😊",
      "第2周·未聊(主聊日8-30)"),
 22: ("周频率·聊天日①","全天可聊(约定)","周频率·主聊",
      "今天周日，咱们这周好好聊会儿～定好闹钟，聊完我得休息😊",
      "来啦，今天约好聊的～",
      "第2周·主聊天日"),
 23: ("周频率·聊天日①","全天可聊(约定)","周频率·主聊",
      "新的一周开始，今天有空咱们聊会儿～定闹钟哈😊",
      "来啦，这周约好的～",
      "第3周(8-31~9-6)·主聊天日"),
}

for rr, (mode, win, exc, main, t17, remark) in N.items():
    if mode.startswith("周频率·聊天日"):
        mcolor = BLUE
    elif mode.startswith("周频率·可选"):
        mcolor = GOLD
    else:
        mcolor = RED
    setcell(rr, 4, mode, color=mcolor, bold=True, center=True)
    setcell(rr, 5, win, color=mcolor, center=True)
    setcell(rr, 6, exc, color=mcolor)
    setcell(rr, 7, main, color="1F1F1F")
    setcell(rr, 8, t17, color="1F1F1F")
    setcell(rr, 9, remark, color="595959")
    ws.cell(rr,1).font = Font(bold=True, size=10, name=F, color="1F1F1F")
    ws.cell(rr,3).font = Font(size=10, name=F, color="1F1F1F")

# ---------- 3) 末尾追加「周频率聊天日总览」 ----------
def section(row, text, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, row, value=text) if False else ws.cell(row, 1, value=text)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=11, name=F)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

r = ws.max_row + 2
section(r, "📅 周频率聊天日总览（8-20 起 · 自然周 周一~周日 · 每周1次为主、最多2次）"); r += 1
# 表头
hdr = ["周次","日期范围","主聊天日①","可选加次②","说明"]
for j, h in enumerate(hdr, start=1):
    c = ws.cell(r, j, value=h)
    c.fill = PatternFill("solid", fgColor=GREY)
    c.font = Font(bold=True, color="FFFFFF", size=10, name=F)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
ws.row_dimensions[r].height = 28
r += 1
rows = [
 ("第1周","8-17 ~ 8-23","8-20(周四)","8-23(周日)","新规8-20生效，本周已用1次（隔离见效）"),
 ("第2周","8-24 ~ 8-30","8-30(周日)","8-27(周四)","每周1次为主；如需第2次选周四"),
 ("第3周","8-31 ~ 9-6","8-31(周一)","—","进入新周，主聊日自行定（示例放周一）"),
]
for wk_range in rows:
    for j, v in enumerate(wk_range, start=1):
        c = ws.cell(r, j, value=v)
        col = BLUE if j in (3,) else (GOLD if j==4 else "1F1F1F")
        c.font = Font(bold=(j in (1,3)), size=10, name=F, color=col)
        c.alignment = Alignment(horizontal=("center" if j in (1,2,3,4) else "left"),
                                vertical="top", wrap_text=True); c.border = border
    ws.row_dimensions[r].height = 30
    r += 1

try:
    wb.save(PATH)
    print("OK saved. sheets:", wb.sheetnames, "| 8-20起已改周频率 | max_row:", ws.max_row)
except PermissionError as e:
    print("PERMISSION_ERROR: 文件被占用，请关闭WPS后重试。", e)
